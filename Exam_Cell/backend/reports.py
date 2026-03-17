"""
Report generation — PDF (ReportLab) and Excel (OpenPyXL) absentee reports.
"""

from __future__ import annotations

import io
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ═══════════════════════════════════════════════════════════════════════════════
# PDF GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

def generate_pdf(
    absentees: list[dict[str, str]],
    exam_info: dict[str, Any],
    report_type: str,
) -> bytes:
    """Generate a PDF absentee report and return raw bytes."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontSize=11,
        alignment=TA_CENTER,
        spaceAfter=14,
        textColor=colors.grey,
    )
    section_style = ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading2"],
        fontSize=13,
        spaceBefore=16,
        spaceAfter=8,
        textColor=colors.HexColor("#2c3e50"),
    )

    elements: list[Any] = []

    # ── Title ──────────────────────────────────────────────────────────────
    elements.append(Paragraph("Absentee Report", title_style))
    exam_str = f"{exam_info.get('date', '')} — {exam_info.get('type', '')}"
    if exam_info.get("session"):
        exam_str += f" ({exam_info['session']})"
    elements.append(Paragraph(exam_str, subtitle_style))
    elements.append(Spacer(1, 0.3 * cm))

    if not absentees:
        elements.append(Paragraph("No absentees recorded.", styles["Normal"]))
        doc.build(elements)
        return buf.getvalue()

    # ── Group data ─────────────────────────────────────────────────────────
    if report_type == "hall":
        groups = _group_by(absentees, "hall_number")
        group_label = "Hall"
    elif report_type == "department":
        groups = _group_by(absentees, "department")
        group_label = "Department"
    elif report_type == "class":
        groups = _group_by(absentees, "class_name")
        group_label = "Class"
    else:
        groups = {"All Students": absentees}
        group_label = "Overall"

    # ── Build tables per group ─────────────────────────────────────────────
    for group_name, students in sorted(groups.items()):
        elements.append(
            Paragraph(f"{group_label}: {group_name}  ({len(students)} absentees)", section_style)
        )

        header = ["#", "Register No", "Student Name", "Class", "Side", "Hall", "Dept"]
        data = [header]
        for idx, st in enumerate(students, 1):
            data.append([
                str(idx),
                st["register_number"],
                st["student_name"],
                st.get("class_name", "-"),
                st.get("side_of_seat", "-"),
                st["hall_number"],
                st["department"],
            ])

        col_widths = [0.8 * cm, 3.2 * cm, 4.5 * cm, 3.0 * cm, 2.5 * cm, 1.8 * cm, 2.7 * cm]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.5 * cm))

    doc.build(elements)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

def generate_excel(
    absentees: list[dict[str, str]],
    exam_info: dict[str, Any],
    report_type: str,
) -> bytes:
    """Generate an XLSX absentee report and return raw bytes."""
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    if report_type == "hall":
        groups = _group_by(absentees, "hall_number")
    elif report_type == "department":
        groups = _group_by(absentees, "department")
    elif report_type == "class":
        groups = _group_by(absentees, "class_name")
    else:
        groups = {"All Absentees": absentees}

    first_sheet = True
    for group_name, students in sorted(groups.items()):
        if first_sheet:
            ws = wb.active
            ws.title = str(group_name)[:31]
            first_sheet = False
        else:
            ws = wb.create_sheet(title=str(group_name)[:31])

        # Title row
        exam_str = f"Absentee Report — {exam_info.get('date', '')} — {exam_info.get('type', '')}"
        if exam_info.get("session"):
            exam_str += f" ({exam_info['session']})"
        ws.merge_cells("A1:G1")
        ws["A1"] = exam_str
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Group: {group_name}  |  Absentees: {len(students)}"
        ws["A2"].font = Font(italic=True, size=10, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")

        # Headers in row 4
        headers = ["#", "Register Number", "Student Name", "Class", "Side",
                    "Department", "Year", "Hall Number"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
# 18600
        # Data rows
        for row_idx, st in enumerate(students, 5):
            ws.cell(row=row_idx, column=1, value=row_idx - 4).border = thin_border
            ws.cell(row=row_idx, column=2, value=st["register_number"]).border = thin_border
            ws.cell(row=row_idx, column=3, value=st["student_name"]).border = thin_border
            ws.cell(row=row_idx, column=4, value=st.get("class_name", "-")).border = thin_border
            ws.cell(row=row_idx, column=5, value=st.get("side_of_seat", "-")).border = thin_border
            ws.cell(row=row_idx, column=6, value=st["department"]).border = thin_border
            ws.cell(row=row_idx, column=7, value=st["year"]).border = thin_border
            ws.cell(row=row_idx, column=8, value=st["hall_number"]).border = thin_border

        # Column widths
        widths = [5, 20, 30, 15, 20, 8, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _group_by(items: list[dict], key: str) -> dict[str, list[dict]]:
    groups: dict[str, list[dict]] = {}
    for item in items:
        k = item.get(key, "Unknown")
        groups.setdefault(k, []).append(item)
    return groups
